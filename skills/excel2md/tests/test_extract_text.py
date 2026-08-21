import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

SKILL_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_DIR))

import xlsx_extract_text  # noqa: E402

SAMPLE = Path(
    r"D:\claude-code\226-vi\BD_266_製造指図 購買発注の日付を合わせる_v1.09_VN.xlsx"
)


def make_wb(path: Path) -> None:
    """3 sheets: visible / hidden / visible, sheet 3 has a strikethrough cell."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Bìa"
    ws1["A1"] = "Lịch sử chỉnh sửa"

    ws2 = wb.create_sheet("Ẩn")
    ws2.sheet_state = "hidden"
    ws2["A1"] = "khong duoc xuat hien"

    ws3 = wb.create_sheet("Bảng")
    ws3["A1"] = "giữ lại"
    ws3["B1"] = "bỏ đi"
    ws3["B1"].font = Font(strike=True)

    wb.save(path)


def test_writes_one_file_pair_per_visible_sheet(tmp_path):
    src = tmp_path / "t.xlsx"
    make_wb(src)
    out = tmp_path / "out"

    xlsx_extract_text.convert(str(src), str(out))

    assert (out / "normal_001_Bìa.txt").exists()
    assert (out / "strike_001_Bìa.txt").exists()
    assert (out / "normal_003_Bảng.txt").exists()
    assert (out / "strike_003_Bảng.txt").exists()


def test_hidden_sheet_produces_no_file(tmp_path):
    src = tmp_path / "t.xlsx"
    make_wb(src)
    out = tmp_path / "out"

    xlsx_extract_text.convert(str(src), str(out))

    assert not (out / "normal_002_Ẩn.txt").exists()
    assert not any(p.name.startswith("normal_002") for p in out.iterdir())


def test_index_counts_hidden_sheets(tmp_path):
    """Sheet thứ 3 giữ index 003 dù sheet 2 bị ẩn."""
    src = tmp_path / "t.xlsx"
    make_wb(src)
    out = tmp_path / "out"

    xlsx_extract_text.convert(str(src), str(out))

    assert (out / "normal_003_Bảng.txt").exists()
    assert not (out / "normal_002_Bảng.txt").exists()


def test_no_legacy_combined_files(tmp_path):
    src = tmp_path / "t.xlsx"
    make_wb(src)
    out = tmp_path / "out"

    xlsx_extract_text.convert(str(src), str(out))

    assert not (out / "normal.txt").exists()
    assert not (out / "strike.txt").exists()


def test_preserves_vietnamese_diacritics(tmp_path):
    src = tmp_path / "t.xlsx"
    make_wb(src)
    out = tmp_path / "out"

    xlsx_extract_text.convert(str(src), str(out))

    text = (out / "normal_001_Bìa.txt").read_text(encoding="utf-8")
    assert "Lịch sử chỉnh sửa" in text


def test_separates_strikethrough_from_normal(tmp_path):
    src = tmp_path / "t.xlsx"
    make_wb(src)
    out = tmp_path / "out"

    xlsx_extract_text.convert(str(src), str(out))

    normal = (out / "normal_003_Bảng.txt").read_text(encoding="utf-8").strip()
    strike = (out / "strike_003_Bảng.txt").read_text(encoding="utf-8").strip()
    assert normal == "giữ lại"
    assert strike == "bỏ đi"


@pytest.mark.skipif(not SAMPLE.exists(), reason="sample workbook not available")
def test_formula_cells_render_as_cached_values(tmp_path):
    out = tmp_path / "out"

    xlsx_extract_text.convert(str(SAMPLE), str(out))

    text = (out / "normal_002_Lịch sử chỉnh sửa.txt").read_text(encoding="utf-8")
    assert "=ROW()" not in text
    assert "=IF(" not in text
    assert "Bổ sung item đồng bộ" in text
