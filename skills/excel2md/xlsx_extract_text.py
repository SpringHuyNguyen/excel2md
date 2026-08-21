"""
Extract text from Excel sheets into per-sheet output files:
  - normal_{IDX:03}_{SHEET}.txt : text without strikethrough
  - strike_{IDX:03}_{SHEET}.txt : text with strikethrough

One pair per visible sheet. {IDX:03} is the sheet's zero-padded index,
counting hidden sheets, so indexes may skip numbers. Hidden sheets
produce no output files.

Usage:
    python xlsx_extract_text.py input.xlsx output_dir
"""

import io
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock


def extract_cell(cell) -> tuple[str, str]:
    """
    Returns (normal_text, strike_text) from a cell.
    Both can be empty strings.
    """
    if cell.value is None:
        return "", ""

    # Entire cell is strikethrough
    if cell.font and cell.font.strike:
        return "", str(cell.value).strip()

    # Rich text: mixed formatting
    if isinstance(cell.value, CellRichText):
        normal_parts = []
        strike_parts = []
        for block in cell.value:
            if isinstance(block, TextBlock):
                text = block.text  # correct attribute
                if block.font and block.font.strike:
                    strike_parts.append(text)
                else:
                    normal_parts.append(text)
            else:
                # plain string segment (no special font)
                normal_parts.append(str(block))
        return "".join(normal_parts).strip(), "".join(strike_parts).strip()

    # Plain cell, no strikethrough
    return str(cell.value).strip(), ""


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name)


def convert(input_path: str, output_dir: str):
    os.makedirs(output_dir, exist_ok=True)
    # rich_text=True để tách được strikethrough; data_only=True để lấy giá trị
    # cached của công thức thay vì chuỗi "=ROW()-5". Hai cờ này kết hợp được.
    wb = load_workbook(input_path, rich_text=True, data_only=True)

    written = 0
    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        # Skip hidden / veryHidden sheets — the customer hid them on purpose.
        if ws.sheet_state != "visible":
            print(f"- '{sheet_name}': skipped (hidden)")
            continue

        sheet_normal = []
        sheet_strike = []

        for row in ws.iter_rows():
            row_normal = []
            row_strike = []
            for cell in row:
                n, s = extract_cell(cell)
                if n:
                    row_normal.append(n)
                if s:
                    row_strike.append(s)
            if row_normal:
                sheet_normal.append("\t".join(row_normal))
            if row_strike:
                sheet_strike.append("\t".join(row_strike))

        prefix = f"{idx:03}_{sanitize_filename(sheet_name)}"
        normal_path = os.path.join(output_dir, f"normal_{prefix}.txt")
        strike_path = os.path.join(output_dir, f"strike_{prefix}.txt")

        # Luôn ghi cả 2 file kể cả rỗng, để đường dẫn truyền cho sub-agent
        # lúc nào cũng tồn tại.
        with open(normal_path, "w", encoding="utf-8") as f:
            f.write("\n".join(sheet_normal))
        with open(strike_path, "w", encoding="utf-8") as f:
            f.write("\n".join(sheet_strike))

        written += 1
        print(f"✓ '{sheet_name}': {len(sheet_normal)} normal rows, "
              f"{len(sheet_strike)} strike rows → {os.path.basename(normal_path)}")

    print(f"\n→ {written} sheet(s) written to {output_dir}")


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    if len(sys.argv) < 3:
        print("Usage: python xlsx_extract_text.py <input.xlsx> <output_dir>")
        sys.exit(1)

    convert(sys.argv[1], sys.argv[2])
